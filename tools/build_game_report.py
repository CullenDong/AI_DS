"""Build the Top-20 report Excel from game_id_filter.csv WITH online research.

Per user:
  - all 供應商 rows ranked together (orders_* not separated; a 源标记 column is
    added for transparency only).
  - deliverable = cleaned full sheet + Top20 + 说明, with 主力市场 / 上线年 /
    波动 / 最终分 from online research (provider 官网 + reputable sources).

Operator market = 大中华/亚洲 facing (currency CNY), so the 市场契合分 rubric is
scored relative to a Greater-China / Asia audience.

最终分 = 归一化( 0.55×运营分 + 0.45×市场契合分 ), 缩放使最高=100
运营分    = mean(已验证赚钱[GGR], 稳定[莊家優勢], 覆盖[玩家數])  (normalized)
市场契合分 = mean(题材契合, 波动偏好, 供应商接受度)
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "input" / "ivi_game_all.csv"
OUT = ROOT / "data" / "output" / "ivi_game_all_report.xlsx"
TEXT_COLS = {"供應商", "遊戲名稱"}

# ---------------------------------------------------------------------------
# Online research, keyed by 遊戲名稱 (unique within the Top20).
# market 主力市场 / vol_off 官网波动 / year 上线年 / gtype 类型
# ---------------------------------------------------------------------------
RESEARCH = {
    "超级幸运7": dict(market="大中华/亚洲", vol_off="N/A(真人桌游)", year="待补", gtype="live"),
    "Jackpot时时奖": dict(market="大中华/亚洲", vol_off="中(即开/街机)", year="待补", gtype="lottery"),
    "Soccer": dict(market="大中华/亚洲/东南亚", vol_off="N/A(体育投注)", year="~1998", gtype="sports"),
    "Basketball": dict(market="大中华/亚洲", vol_off="N/A(体育投注)", year="待补", gtype="sports"),
    "庄保险": dict(market="大中华/亚洲", vol_off="N/A(真人桌游)", year="待补", gtype="live"),
    "VIP竞咪百家乐": dict(market="大中华/亚洲(VIP)", vol_off="N/A(真人桌游)", year="待补", gtype="live"),
    "百家乐": dict(market="全球(含亚洲)", vol_off="N/A(真人桌游)", year="~2006", gtype="live"),
    "Combo": dict(market="大中华/亚洲/东南亚", vol_off="N/A(体育投注)", year="待补", gtype="sports"),
    "50倍场": dict(market="大中华/东南亚", vol_off="中(街机捕鱼)", year="待补", gtype="fishing"),
    "大": dict(market="大中华/亚洲", vol_off="N/A(真人桌游)", year="待补", gtype="live"),
    "CyberSoccer": dict(market="大中华/亚洲/东南亚", vol_off="N/A(电竞投注)", year="待补", gtype="esports"),
    "CyberBasketball": dict(market="大中华/亚洲/东南亚", vol_off="N/A(电竞投注)", year="待补", gtype="esports"),
    "超和5": dict(market="大中华/亚洲", vol_off="待补", year="待补", gtype="lottery"),
    "财神捕鱼": dict(market="大中华/东南亚", vol_off="中(街机捕鱼)", year="~2020", gtype="fishing"),
    "Tennis": dict(market="亚洲/全球", vol_off="N/A(体育投注)", year="待补", gtype="sports"),
    "Time of First Goal": dict(market="大中华/亚洲/东南亚", vol_off="N/A(体育投注)", year="待补", gtype="sports"),
    "麻将胡了": dict(market="大中华/亚洲/东南亚", vol_off="Medium", year="~2019", gtype="slot"),
    "麻将胡了2": dict(market="大中华/亚洲/东南亚", vol_off="Medium", year="~2020", gtype="slot"),
    "老虎对": dict(market="大中华/亚洲", vol_off="N/A(真人桌游)", year="~2024", gtype="live"),
    "自动下注": dict(market="亚洲", vol_off="待补", year="~2017", gtype="other"),
    "串关": dict(market="大中华/亚洲/东南亚", vol_off="N/A(体育投注)", year="待补", gtype="sports"),
}

# market-fit rubric (0..1), operator market = 大中华/亚洲 -----------------------
PROVIDER_ACCEPT = {  # 供应商接受度 (base provider, orders_ stripped)
    "SHABA": 1.00, "AGQJ": 0.95, "AGIN": 0.90, "EVO": 0.90, "JDB": 0.90, "FISH": 0.85,
    "PG": 0.90, "YOPLAY": 0.80,
}
VOL_PREF_MAP = {"Low": 0.6, "Medium-Low": 0.8, "Medium": 1.0, "Medium-High": 1.0, "High": 0.8, "Very High": 0.6}


def theme_fit(market: str) -> float:
    if "大中华" in market:
        return 1.00
    if "亚洲" in market:
        return 0.95
    if "东南亚" in market:
        return 0.90
    if "全球" in market:
        return 0.70
    return 0.60


def vol_pref_score(vol_off: str) -> float:
    if vol_off.startswith("N/A") or vol_off == "待补":
        return 0.80  # neutral for sports/table/lottery or unknown
    if vol_off.startswith("中"):
        return 1.00  # fishing/arcade medium fits the casual Asian market well
    return VOL_PREF_MAP.get(vol_off, 0.80)


def data_vol_tier(max_mult: float, gtype: str) -> str:
    if gtype != "slot":
        return "N/A(非slot)"
    if pd.isna(max_mult):
        return "—"
    if max_mult < 50:
        return "Low"
    if max_mult < 200:
        return "Medium-Low"
    if max_mult < 1000:
        return "Medium"
    if max_mult < 3000:
        return "Medium-High"
    if max_mult < 8000:
        return "High"
    return "Very High"


def reconcile_vol(data_vol: str, vol_off: str) -> str:
    if data_vol.startswith("N/A"):
        return vol_off
    if vol_off.startswith("N/A") or vol_off == "待补" or data_vol == vol_off:
        return data_vol
    return f"{data_vol}(官网{vol_off})"


def norm(s: pd.Series) -> pd.Series:
    lo, hi = s.min(), s.max()
    if hi == lo:
        return pd.Series([0.5] * len(s), index=s.index)
    return (s - lo) / (hi - lo)


def main() -> None:
    df = pd.read_csv(SRC)
    for c in df.columns:
        if c not in TEXT_COLS:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    df["RTP%"] = (100 - df["莊家優勢%"]).round(3)

    top = df.sort_values("總投注萬_CNY", ascending=False).head(20).reset_index(drop=True)
    top["运营排名"] = top.index + 1  # ranking by total bet

    rows = []
    for _, r in top.iterrows():
        game = r["遊戲名稱"]
        base = str(r["供應商"]).replace("orders_", "").upper()
        info = RESEARCH.get(game, dict(market="待补", vol_off="待补", year="待补", gtype="other"))
        d_vol = data_vol_tier(r["最大倍率"], info["gtype"])
        rows.append(dict(
            运营排名=int(r["运营排名"]), 游戏=game, 供应商=r["供應商"], base=base,
            上线年=info["year"], 主力市场=info["market"], 类型=info["gtype"],
            RTP=r["RTP%"], 波动=reconcile_vol(d_vol, info["vol_off"]),
            波动_数据=d_vol, 波动_官网=info["vol_off"],
            GGR万=round(r["GGR萬_CNY"], 1), 总投注万=round(r["總投注萬_CNY"], 1),
            玩家數=int(r["玩家數"]), 莊家優勢=round(r["莊家優勢%"], 3),
            最大倍率=round(r["最大倍率"], 1), vol_off=info["vol_off"], market=info["market"],
        ))
    t = pd.DataFrame(rows)

    # ---- scoring -----------------------------------------------------------
    earn = norm(t["GGR万"])                  # 已验证赚钱
    stable = norm(t["莊家優勢"])              # 稳定
    reach = norm(t["玩家數"].astype(float))   # 覆盖
    t["运营分"] = ((earn + stable + reach) / 3 * 100).round(1)

    t["题材分"] = [theme_fit(m) for m in t["market"]]
    t["波动偏好分"] = [vol_pref_score(v) for v in t["vol_off"]]
    t["供应商接受度分"] = [PROVIDER_ACCEPT.get(b, 0.70) for b in t["base"]]
    t["市场契合分"] = ((t["题材分"] + t["波动偏好分"] + t["供应商接受度分"]) / 3 * 100).round(1)

    raw = 0.55 * t["运营分"] + 0.45 * t["市场契合分"]
    t["最终分"] = (raw / raw.max() * 100).round(1)
    t = t.sort_values("最终分", ascending=False).reset_index(drop=True)
    t["最终排名"] = t.index + 1
    t["Δ"] = t["运营排名"] - t["最终排名"]
    t["Δ显示"] = t["Δ"].apply(lambda x: "—" if x == 0 else (f"↑{x}" if x > 0 else f"↓{abs(x)}"))

    show = t[[
        "最终排名", "游戏", "供应商", "上线年", "主力市场", "RTP", "波动",
        "GGR万", "玩家數", "运营分", "市场契合分", "最终分", "Δ显示", "运营排名",
        "波动_数据", "波动_官网", "总投注万", "莊家優勢",
    ]].rename(columns={"最终排名": "#", "Δ显示": "Δ", "RTP": "RTP%", "莊家優勢": "莊家優勢%"})

    notes = pd.DataFrame({
        "项": [
            "数据来源", "操盘市场假设", "研究范围", "口径处理", "排名依据",
            "最终分公式", "运营分(55%)", "市场契合分(45%)", "主力市场/上线年", "波动",
            "RTP%", "供应商接受度", "题材契合", "波动偏好", "待补项",
        ],
        "说明": [
            "data/input/ivi_game_all.csv (繁体表头, 单位 萬_CNY) + 各游戏官网/权威评测联网研究",
            "按 CNY 推断为 大中华/亚洲 向, 市场契合分据此评估",
            "全量 4506 行清洗; Top20 取自 總投注 最高 20 个; 联网研究其全部唯一游戏",
            "本文件已为合并/去重版本(无 orders_ 双口径), 直接排名",
            "Top20 选取按 總投注; 重排后按 最终分 排序, Δ=运营排名−最终排名",
            "最终分 = 归一化(0.55×运营分 + 0.45×市场契合分), 缩放使最高=100",
            "mean(已验证赚钱[GGR归一], 稳定[莊家優勢归一], 覆盖[玩家數归一]) ×100",
            "mean(题材契合, 波动偏好, 供应商接受度) ×100",
            "联网查厂商官网/权威源: SHABA=沙巴体育SABA, AGQJ/AGIN=AG真人(Asia Gaming), JDB捕鱼, EVO=Evolution",
            "本批多为体育投注/真人桌游/捕鱼, 非slot波动概念不适用→标N/A并用官网口径; slot才用数据档(最大倍率)",
            "由 100% − 莊家優勢% 反推 (原表无 RTP 列)",
            "SHABA1.00/AGQJ0.95/AGIN·EVO·JDB0.90/FISH0.85",
            "大中华1.0 / 亚洲0.95 / 东南亚0.9 / 全球0.7",
            "Medium·MH=1.0 / 中(街机)=1.0 / ML·High=0.8 / N/A·待补=0.8(中性) / Low·VH=0.6",
            "AGIN(Jackpot时时奖/超和5)等小厂上线年/部分波动官网无公开数据, 标“待补”",
        ],
    })

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUT, engine="openpyxl") as xw:
        show.to_excel(xw, sheet_name="Top20重排", index=False)
        notes.to_excel(xw, sheet_name="说明", index=False)
        df.to_excel(xw, sheet_name="全量数据(4506)", index=False)
        _format(xw, show)

    print(f"written: {OUT}")
    print(show[["#", "游戏", "供应商", "主力市场", "波动", "最终分", "Δ"]].to_string(index=False))


def _format(xw, show):
    wb = xw.book
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws = wb["Top20重排"]
    for j, col in enumerate(show.columns, 1):
        c = ws.cell(1, j)
        c.fill, c.font, c.alignment = header_fill, header_font, center
        width = max(9, min(24, int(show[col].astype(str).map(len).max()) + 3))
        ws.column_dimensions[get_column_letter(j)].width = width
    ws.freeze_panes = "C2"
    ws.row_dimensions[1].height = 30
    fill_alt = PatternFill("solid", fgColor="EAF1FB")
    for i in range(2, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            ws.cell(i, j).alignment = Alignment(horizontal="center", vertical="center")
        if i % 2 == 0:
            for j in range(1, ws.max_column + 1):
                ws.cell(i, j).fill = fill_alt

    wsn = wb["说明"]
    wsn.column_dimensions["A"].width = 16
    wsn.column_dimensions["B"].width = 95
    for j in range(1, 3):
        wsn.cell(1, j).fill = header_fill
        wsn.cell(1, j).font = header_font
    for i in range(2, wsn.max_row + 1):
        wsn.cell(i, 2).alignment = Alignment(wrap_text=True, vertical="top")


if __name__ == "__main__":
    main()
