"""Build the Top-20 casino game report Excel from casinoplus CSV.

Pipeline:
  1. Load & clean the raw CSV (Chinese headers, numbers stored as strings).
  2. Rank by 总投注 (TOTAL_BET) and take Top 20 -> this is the "运营排序".
  3. Derive a data-driven volatility tier from 最大倍率.
  4. Merge online research (主力市场 / 官网波动 / 上线年 / 题材) per game.
  5. Score each game per the image formula:
       最终分 = norm( 0.55*运营分 + 0.45*市场契合分 ), scaled so max = 100
     运营分   = mean(已验证赚钱[GGR], 稳定[庄家优势], 覆盖[玩家数])   (normalized)
     市场契合分 = mean(题材, 波动偏好, 供应商接受度)
  6. Δ = rank change between the operations ranking and the re-ranked final ranking.
  7. Write a formatted multi-sheet Excel to data/output/.

All scoring assumptions live in the dicts below and in the 评分说明 sheet so they
are easy to tweak.
"""
from __future__ import annotations

import math
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "data" / "input" / "casinoplus_2024_01-06.csv"
OUT = ROOT / "data" / "output" / "casinoplus_top20_report.xlsx"
PERIOD_MONTHS = 6  # data covers 2024-01..2024-06

# ---------------------------------------------------------------------------
# Online research (provider official sites + reputable review sources).
# keyed by (provider, game). See chat / 评分说明 sheet for source URLs.
# ---------------------------------------------------------------------------
RESEARCH = {
    ("JILI", "Super Ace"): dict(market="菲律宾/东南亚(拉美增长)", vol_off="Medium-High", year="~2021", gtype="slot"),
    ("JILI", "Fortune Gems"): dict(market="菲律宾/东南亚", vol_off="Medium-Low", year="~2021", gtype="slot"),
    ("EVOLUTION", "Speed Baccarat A"): dict(market="全球(欧亚拉美)", vol_off="N/A(桌游)", year="~2017", gtype="live"),
    ("COLORGAME", "livecolorgame"): dict(market="菲律宾(perya色彩游戏)", vol_off="N/A(桌游)", year="~2024", gtype="live"),
    ("JILI", "Golden Empire"): dict(market="菲律宾/东南亚", vol_off="Medium", year="~2021", gtype="slot"),
    ("JILI", "Money Coming"): dict(market="菲律宾/东南亚(印度)", vol_off="High", year="~2021", gtype="slot"),
    ("JILI", "Mines"): dict(market="菲律宾/东南亚(印度)", vol_off="Medium-High", year="~2023", gtype="arcade"),
    ("TIPTOP", "Pusoy"): dict(market="菲律宾(本土纸牌)", vol_off="N/A(纸牌)", year="~2019", gtype="card"),
    ("EVOLUTION", "Crazy Time"): dict(market="全球(欧/拉美/亚)", vol_off="High", year="~2020", gtype="live"),
    ("TIPTOP", "Tongits Go"): dict(market="菲律宾(本土纸牌)", vol_off="N/A(纸牌)", year="~2019", gtype="card"),
    ("JILI", "Fortune Gems 2"): dict(market="菲律宾/东南亚", vol_off="Medium", year="2023", gtype="slot"),
    ("JILI", "Tongits Star"): dict(market="菲律宾(本土纸牌)", vol_off="N/A(纸牌)", year="~2022", gtype="card"),
    ("JILI", "Fortune Gems 3"): dict(market="菲律宾/东南亚", vol_off="Medium-High", year="~2024", gtype="slot"),
    ("JILI", "Boxing King"): dict(market="菲律宾/东南亚", vol_off="Medium", year="2021", gtype="slot"),
    ("JILI", "Crazy777"): dict(market="菲律宾/东南亚", vol_off="Medium-Low", year="2021", gtype="slot"),
    ("EVOLUTION", "No Commission Baccarat"): dict(market="全球(强亚洲/大中华)", vol_off="N/A(桌游)", year="~2017", gtype="live"),
    ("EVOLUTION", "Golden Wealth Baccarat"): dict(market="全球(强亚洲/大中华)", vol_off="N/A(桌游)", year="~2021", gtype="live"),
    ("PG", "Baccarat Deluxe"): dict(market="东南亚/大中华", vol_off="Low", year="~2019", gtype="card"),
    ("JILI", "Wild Ace"): dict(market="菲律宾/东南亚", vol_off="Medium", year="~2023", gtype="slot"),
    ("TIPTOP", "Mines"): dict(market="菲律宾(JILI引擎)", vol_off="Medium", year="~2023", gtype="arcade"),
}

# market-fit rubric (all 0..1) -------------------------------------------------
PROVIDER_ACCEPT = {  # 供应商接受度 for the PH market
    "JILI": 1.00, "TIPTOP": 0.95, "COLORGAME": 0.90, "PG": 0.90, "EVOLUTION": 0.85,
}
VOL_PREF = {  # 波动偏好: assume PH casual market favours Medium / Medium-High
    "Low": 0.60, "Medium-Low": 0.80, "Medium": 1.00, "Medium-High": 1.00,
    "High": 0.80, "Very High": 0.60,
}


def theme_fit(market: str) -> float:
    """题材契合度 by primary market string."""
    if "菲律宾" in market:
        return 1.00
    if "东南亚" in market:
        return 0.90
    if "大中华" in market or "亚洲" in market:
        return 0.80
    return 0.60  # 全球/欧美拉美


def vol_pref_score(vol_off: str, data_vol: str) -> float:
    if vol_off.startswith("N/A"):
        return 0.80  # table/card games: strong but distinct appeal in PH -> neutral-high
    return VOL_PREF.get(vol_off, VOL_PREF.get(data_vol, 0.8))


def data_vol_tier(max_mult: float, gtype: str) -> str:
    """Data-driven volatility tier from observed 最大倍率."""
    if gtype in ("live", "card"):
        return "N/A(非slot)"
    if max_mult < 50:
        return "Low"
    if max_mult < 200:
        return "Medium-Low"
    if max_mult < 1000:
        return "Medium"
    if max_mult < 3000:
        return "Medium-High"
    if max_mult < 8000:
        return "High"
    return "Very High"


def reconcile_vol(data_vol: str, vol_off: str) -> str:
    """波动 = 数据算为主, 官网补充. Show both when they diverge."""
    if data_vol.startswith("N/A"):
        return vol_off  # rely on official for non-slots
    if vol_off.startswith("N/A") or data_vol == vol_off:
        return data_vol
    return f"{data_vol}(官网{vol_off})"


def norm(s: pd.Series) -> pd.Series:
    lo, hi = s.min(), s.max()
    if hi == lo:
        return pd.Series([0.5] * len(s), index=s.index)
    return (s - lo) / (hi - lo)


def main() -> None:
    df = pd.read_csv(SRC)
    # numeric coercion (everything but the two text cols)
    text_cols = {"平台ID", "下单平台", "游戏名称"}
    for c in df.columns:
        if c not in text_cols:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    top = df.sort_values("总投注万_PHP", ascending=False).head(20).reset_index(drop=True)
    top["运营排名"] = top.index + 1  # operations ranking (by total bet)

    rows = []
    for _, r in top.iterrows():
        prov, game = r["下单平台"], r["游戏名称"]
        info = RESEARCH.get((prov, game), dict(market="待补", vol_off="待补", year="待补", gtype="slot"))
        d_vol = data_vol_tier(r["最大倍率"], info["gtype"])
        rows.append(
            dict(
                运营排名=int(r["运营排名"]),
                游戏=game,
                商=prov,
                上线年=info["year"],
                主力市场=info["market"],
                RTP=round(r["RTP"] * 100, 2),
                波动=reconcile_vol(d_vol, info["vol_off"]),
                波动_数据=d_vol,
                波动_官网=info["vol_off"],
                累计GGR万=round(r["GGR万_PHP"], 1),
                总投注万=round(r["总投注万_PHP"], 1),
                玩家数=int(r["玩家数"]),
                月均玩家万=round(r["玩家数"] / PERIOD_MONTHS / 1e4, 2),
                庄家优势pct=round(r["庄家优势_pct"], 3),
                gtype=info["gtype"],
                market=info["market"],
                vol_off=info["vol_off"],
            )
        )
    t = pd.DataFrame(rows)

    # ---- scoring -----------------------------------------------------------
    earn = norm(t["累计GGR万"])           # 已验证赚钱
    stable = norm(t["庄家优势pct"])        # 稳定 (reliable per-bet margin)
    reach = norm(t["玩家数"].astype(float))  # 覆盖
    t["运营分"] = ((earn + stable + reach) / 3 * 100).round(1)

    t["题材分"] = [theme_fit(m) for m in t["market"]]
    t["波动偏好分"] = [vol_pref_score(vo, dv) for vo, dv in zip(t["vol_off"], t["波动_数据"])]
    t["供应商接受度分"] = [PROVIDER_ACCEPT.get(p, 0.7) for p in t["商"]]
    t["市场契合分"] = ((t["题材分"] + t["波动偏好分"] + t["供应商接受度分"]) / 3 * 100).round(1)

    raw = 0.55 * t["运营分"] + 0.45 * t["市场契合分"]
    t["最终分"] = (raw / raw.max() * 100).round(1)

    t = t.sort_values("最终分", ascending=False).reset_index(drop=True)
    t["最终排名"] = t.index + 1
    t["Δ"] = t["运营排名"] - t["最终排名"]  # +升 / -降
    t["Δ显示"] = t["Δ"].apply(lambda x: "—" if x == 0 else (f"↑{x}" if x > 0 else f"↓{abs(x)}"))

    # ---- presentation table (image column order) ---------------------------
    show = t[[
        "最终排名", "游戏", "商", "上线年", "主力市场", "RTP", "波动",
        "累计GGR万", "月均玩家万", "运营分", "市场契合分", "最终分", "Δ显示", "运营排名",
        "波动_数据", "波动_官网", "总投注万", "玩家数", "庄家优势pct",
    ]].rename(columns={
        "最终排名": "#", "商": "供应商", "累计GGR万": "累计GGR(1-6月)万",
        "月均玩家万": "月均玩家(万)", "Δ显示": "Δ", "庄家优势pct": "庄家优势%",
    })

    # ---- methodology sheet -------------------------------------------------
    notes = pd.DataFrame({
        "项": [
            "数据周期", "研究范围", "运营排序依据", "最终分公式", "运营分(55%)",
            "市场契合分(45%)", "波动(数据)", "波动(官网)", "波动(最终)",
            "月均玩家", "累计GGR说明", "Δ", "供应商接受度", "波动偏好(PH)", "题材契合",
        ],
        "说明": [
            "2024-01 ~ 2024-06 (6个月)",
            "按 总投注万_PHP 取 Top 20",
            "按 总投注 降序 = 上一轮“运营排序”",
            "最终分 = 归一化(0.55×运营分 + 0.45×市场契合分), 缩放使最高=100",
            "mean(已验证赚钱[GGR归一], 稳定[庄家优势归一], 覆盖[玩家数归一]) ×100",
            "mean(题材契合, 波动偏好, 供应商接受度) ×100",
            "由观测 最大倍率 分档: <50 Low / <200 ML / <1000 M / <3000 MH / <8000 High / ≥8000 VeryHigh",
            "来自厂商官网/权威评测 (JILI官方多标注1-3/5; PG/Evolution官网)",
            "数据为主、官网补充; 二者不一致时显示“数据(官网X)”",
            "玩家数 ÷ 6个月 ÷ 1万 (玩家数为周期累计口径, 非去重月活)",
            "原图为“三月GGR”; 本表无月度拆分, 用1-6月累计GGR, 已重命名列名标注",
            "运营排名 − 最终排名; +为重排后上升, −为下降",
            "JILI1.00/TIPTOP0.95/COLORGAME0.90/PG0.90/EVOLUTION0.85",
            "Medium·Medium-High=1.0 / ML·High=0.8 / Low·VeryHigh=0.6 / 桌纸牌=0.8",
            "含菲律宾1.0 / 东南亚0.9 / 大中华·亚洲0.8 / 全球欧美拉美0.6",
        ],
    })

    # ---- write Excel -------------------------------------------------------
    OUT.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUT, engine="openpyxl") as xw:
        show.to_excel(xw, sheet_name="Top20重排", index=False)
        notes.to_excel(xw, sheet_name="评分说明", index=False)
        df_clean = df.copy()
        df_clean.to_excel(xw, sheet_name="全量数据(1344)", index=False)
        _format(xw, show, notes)

    print(f"written: {OUT}")
    print(show[["#", "游戏", "供应商", "主力市场", "波动", "最终分", "Δ"]].to_string(index=False))


def _format(xw, show, notes):
    wb = xw.book
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws = wb["Top20重排"]
    for j, col in enumerate(show.columns, 1):
        c = ws.cell(1, j)
        c.fill, c.font, c.alignment = header_fill, header_font, center
        width = max(10, min(26, int(show[col].astype(str).map(len).max()) + 4, 26))
        ws.column_dimensions[get_column_letter(j)].width = width
    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 30
    # zebra + center
    fill_alt = PatternFill("solid", fgColor="EAF1FB")
    for i in range(2, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            ws.cell(i, j).alignment = Alignment(horizontal="center", vertical="center")
        if i % 2 == 0:
            for j in range(1, ws.max_column + 1):
                ws.cell(i, j).fill = fill_alt

    wsn = wb["评分说明"]
    wsn.column_dimensions["A"].width = 16
    wsn.column_dimensions["B"].width = 90
    for j in range(1, 3):
        c = wsn.cell(1, j)
        c.fill, c.font = header_fill, header_font
    for i in range(2, wsn.max_row + 1):
        wsn.cell(i, 2).alignment = Alignment(wrap_text=True, vertical="top")


if __name__ == "__main__":
    main()
